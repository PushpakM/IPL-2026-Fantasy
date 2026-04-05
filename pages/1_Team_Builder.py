import streamlit as st
import pandas as pd
import json
from datetime import date
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from core.data_loader import load_enriched_players, load_schedule, load_venue_data
from core.team_selector import (
    build_my11circle_team, build_my11circle_team_modes, build_tata_ipl_team,
    suggest_differential_picks,
)
from core.transfer_engine import (
    load_state, save_state, set_initial_team, apply_transfers,
    change_captain_vc, recommend_transfers, get_squad_health_report,
)
from core.scraper import scrape_playing_xi, set_playing_xi_manual, get_playing_xi_cached, get_weather_for_match, set_weather_manual
from core.validator import validate_team

st.set_page_config(page_title="Team Builder | IPL 2026", page_icon="🏏", layout="wide")

st.title("Team Builder")


@st.cache_data(ttl=300)
def get_data():
    players = load_enriched_players()
    schedule = load_schedule()
    venues = load_venue_data()
    return players, schedule, venues


players_df, schedule_df, venues_df = get_data()
state = load_state()

# --- Match Selector ---
st.subheader("Select Match")

today = pd.Timestamp(date.today())
upcoming = schedule_df[schedule_df["Date"] >= today].sort_values("Date")
default_idx = 0

match_options = []
for _, row in schedule_df.iterrows():
    match_num = int(row["Match #"])
    match_str = f"Match {match_num}: {row['Match']} — {row['VenueName']}"
    if hasattr(row["Date"], "strftime"):
        match_str += f" ({row['Date'].strftime('%b %d')})"
    match_options.append(match_str)

if len(upcoming) > 0:
    next_match_num = int(upcoming.iloc[0]["Match #"])
    default_idx = next_match_num - 1

selected_idx = st.selectbox(
    "Choose match",
    range(len(match_options)),
    index=min(default_idx, len(match_options) - 1),
    format_func=lambda i: match_options[i],
)

match_row = schedule_df.iloc[selected_idx]
match_info = {
    "Match #": int(match_row["Match #"]),
    "Team1": match_row["Team1"],
    "Team2": match_row["Team2"],
    "VenueName": match_row["VenueName"],
    "Date": match_row["Date"],
    "Match": match_row["Match"],
}

# Get next match for TATA IPL lookahead
next_match_info = None
if selected_idx + 1 < len(schedule_df):
    next_row = schedule_df.iloc[selected_idx + 1]
    next_match_info = {
        "Team1": next_row["Team1"],
        "Team2": next_row["Team2"],
        "VenueName": next_row["VenueName"],
    }

# --- Venue Info ---
venue_match = venues_df[venues_df["Venue"].str.contains(match_info["VenueName"][:15], case=False, na=False)]
if len(venue_match) > 0:
    v = venue_match.iloc[0]
    vcol1, vcol2, vcol3 = st.columns(3)
    with vcol1:
        st.metric("Venue Character", v.get("Batting/Bowling", "Balanced"))
    with vcol2:
        st.metric("Avg 1st Inn Score", int(v.get("Avg 1st Inn Score", 160)))
    with vcol3:
        st.caption(f"**Insight:** {v.get('Key Insight', 'N/A')}")

# --- Weather Conditions ---
st.markdown("---")
st.subheader("Weather & Match Conditions")

weather_data = get_weather_for_match(match_info["VenueName"])

if weather_data:
    wcol1, wcol2, wcol3, wcol4, wcol5 = st.columns(5)
    with wcol1:
        st.metric("Temperature", f"{weather_data.get('temp_c', '--')}°C")
    with wcol2:
        st.metric("Humidity", f"{weather_data.get('humidity', '--')}%")
    with wcol3:
        dew = weather_data.get("dew_factor", 0)
        dew_label = "Heavy" if dew >= 0.7 else "Moderate" if dew >= 0.4 else "Low"
        st.metric("Dew Factor", dew_label, delta=f"{dew:.1f}")
    with wcol4:
        st.metric("Wind", f"{weather_data.get('wind_kph', '--')} kph")
    with wcol5:
        rain = weather_data.get("rain_chance", 0)
        rain_emoji = "🌧" if rain >= 50 else "🌦" if rain >= 25 else "☀️"
        st.metric("Rain Chance", f"{rain_emoji} {rain}%")

    # Fantasy impact summary
    impact = weather_data.get("fantasy_impact", {})
    if impact:
        impact_items = []
        if impact.get("swing_factor", 1.0) > 1.0:
            impact_items.append(f"Swing: +{(impact['swing_factor']-1)*100:.0f}% pace")
        if impact.get("spin_factor", 1.0) > 1.0:
            impact_items.append(f"Spin: +{(impact['spin_factor']-1)*100:.0f}% spinners")
        if impact.get("batting_factor", 1.0) > 1.0:
            impact_items.append(f"Dew: +{(impact['batting_factor']-1)*100:.0f}% batting")
        if impact.get("pace_factor", 1.0) > 1.0:
            impact_items.append(f"Wind: +{(impact['pace_factor']-1)*100:.0f}% pace")

        if impact_items:
            st.info("**Weather Boosts:** " + " | ".join(impact_items))

        summary = impact.get("summary", "")
        if summary:
            st.caption(f"**Analysis:** {summary}")

        rain_risk = impact.get("rain_risk", "none")
        if rain_risk in ("moderate", "high"):
            st.warning(f"**Rain Risk: {rain_risk.upper()}** — Consider high-floor, safe picks. Avoid volatile players in case of DLS.")

    # Source indicator
    source = weather_data.get("source", "defaults")
    if source == "defaults":
        st.caption("Using seasonal defaults. Update with live data below for better accuracy.")

# Manual weather override
with st.expander("Update Weather Manually"):
    st.caption("Enter live weather from your weather app or TV broadcast for best accuracy.")
    mw_col1, mw_col2, mw_col3 = st.columns(3)
    with mw_col1:
        mw_temp = st.number_input("Temperature (°C)", min_value=15, max_value=50, value=weather_data.get("temp_c", 32), key="mw_temp")
        mw_humidity = st.number_input("Humidity (%)", min_value=10, max_value=100, value=weather_data.get("humidity", 55), key="mw_humidity")
    with mw_col2:
        mw_dew = st.slider("Dew Factor (0=none, 1=heavy)", 0.0, 1.0, float(weather_data.get("dew_factor", 0.5)), 0.1, key="mw_dew")
        mw_wind = st.number_input("Wind (kph)", min_value=0, max_value=50, value=weather_data.get("wind_kph", 10), key="mw_wind")
    with mw_col3:
        mw_condition = st.selectbox("Condition", ["Clear", "Partly Cloudy", "Overcast", "Hot & Humid", "Hot & Dry", "Warm", "Humid & Breezy", "Pleasant"], index=0, key="mw_condition")
        mw_rain = st.number_input("Rain Chance (%)", min_value=0, max_value=100, value=weather_data.get("rain_chance", 5), key="mw_rain")

    if st.button("Save Weather Data"):
        updated_weather = set_weather_manual(
            match_info["VenueName"], mw_temp, mw_humidity, mw_dew, mw_wind, mw_condition, mw_rain
        )
        st.success(f"Weather updated! Impact: {updated_weather['fantasy_impact']['summary'][:100]}")
        st.rerun()

st.markdown("---")

# --- Playing XI Section ---
st.subheader("Playing XI (Scrape at Toss Time)")
st.caption("Scrape confirmed playing XI ~30 mins before match. Only confirmed players will be picked.")

xi_col1, xi_col2 = st.columns([2, 1])

with xi_col1:
    if st.button("Scrape Playing XI Now", type="secondary"):
        with st.spinner(f"Scraping playing XI for {match_info['Team1']} vs {match_info['Team2']}..."):
            xi_data = scrape_playing_xi(match_info["Team1"], match_info["Team2"])
            if xi_data:
                st.session_state["playing_xi"] = xi_data
                for key in ("team1", "team2"):
                    td = xi_data.get(key, {})
                    if isinstance(td, dict) and td.get("players"):
                        st.success(f"{td['team']}: {len(td['players'])} players confirmed")
            else:
                st.warning("Could not scrape playing XI. Enter manually or generate without XI filter.")

with xi_col2:
    cached_xi = get_playing_xi_cached()
    if cached_xi:
        st.info("Cached playing XI available")

# Manual playing XI input
with st.expander("Enter Playing XI Manually"):
    manual_col1, manual_col2 = st.columns(2)
    team1_players_available = players_df[players_df["Team"] == match_info["Team1"]]["Player Name"].tolist()
    team2_players_available = players_df[players_df["Team"] == match_info["Team2"]]["Player Name"].tolist()

    with manual_col1:
        st.markdown(f"**{match_info['Team1']}**")
        team1_xi = st.multiselect(f"Select {match_info['Team1']} XI", team1_players_available, max_selections=11, key="manual_xi_team1")
    with manual_col2:
        st.markdown(f"**{match_info['Team2']}**")
        team2_xi = st.multiselect(f"Select {match_info['Team2']} XI", team2_players_available, max_selections=11, key="manual_xi_team2")

    if st.button("Set Manual Playing XI"):
        if team1_xi and team2_xi:
            xi_data = set_playing_xi_manual(match_info["Team1"], match_info["Team2"], team1_xi, team2_xi)
            st.session_state["playing_xi"] = xi_data
            st.success(f"Playing XI set: {match_info['Team1']} ({len(team1_xi)}), {match_info['Team2']} ({len(team2_xi)})")
        else:
            st.warning("Select at least some players for both teams.")

playing_xi = st.session_state.get("playing_xi") or cached_xi

if playing_xi:
    with st.expander("View Current Playing XI", expanded=False):
        for key in ("team1", "team2"):
            team_data = playing_xi.get(key, {})
            if isinstance(team_data, dict) and team_data.get("players"):
                team_label = team_data.get("team", key)
                st.markdown(f"**{team_label}:** {', '.join(team_data['players'])}")
        toss = playing_xi.get("toss", {})
        if toss and toss.get("winner"):
            st.markdown(f"**Toss:** {toss['winner']} won, elected to {toss.get('decision', '?')}")

st.markdown("---")

# =====================================================================
# MY11CIRCLE — Fresh team with risk modes
# =====================================================================
st.subheader("My11Circle — Fresh Team Every Match")
st.caption("No transfers. Build the best XI from scratch for this match.")

risk_mode = st.radio("Strategy Mode", ["Safe", "Balanced", "Aggressive"], index=1, horizontal=True, key="my11_risk")
risk_mode_key = risk_mode.lower()

gen_col1, gen_col2 = st.columns(2)
with gen_col1:
    generate_single = st.button(f"Generate {risk_mode} Team", type="primary", use_container_width=True)
with gen_col2:
    generate_all_modes = st.button("Generate All 3 Modes", use_container_width=True)

if generate_single:
    with st.spinner(f"Building {risk_mode} My11Circle team..."):
        result = build_my11circle_team(
            players_df, match_info["Team1"], match_info["Team2"],
            match_info["VenueName"], playing_xi, risk_mode_key,
        )
        st.session_state["my11circle_team"] = result
        st.session_state.pop("my11circle_modes", None)

if generate_all_modes:
    with st.spinner("Building all 3 My11Circle teams..."):
        modes = build_my11circle_team_modes(
            players_df, match_info["Team1"], match_info["Team2"],
            match_info["VenueName"], playing_xi,
        )
        st.session_state["my11circle_modes"] = modes
        st.session_state.pop("my11circle_team", None)


def display_team(team_data, label):
    """Display a generated team card."""
    if team_data.get("error"):
        st.error(team_data["error"])
        return

    badges = [f"{team_data['credits_used']}/100 cr", f"Est. {team_data['total_points']} pts"]
    if team_data.get("used_playing_xi"):
        badges.append("XI filtered")
    if team_data.get("weather"):
        w = team_data["weather"]
        badges.append(f"{w['temp_c']}°C {w['condition']}")
    if team_data.get("strategy"):
        badges.append(team_data["strategy"][:50])

    if team_data["is_valid"]:
        st.success(" | ".join(badges))
    else:
        st.error("Invalid: " + ", ".join(team_data["validation_errors"]))

    st.markdown(f"**C (2x):** {team_data['captain']} | **VC (1.5x):** {team_data['vice_captain']}")

    rows = []
    for p in team_data["players"]:
        marker = " (C)" if p["Player Name"] == team_data["captain"] else (" (VC)" if p["Player Name"] == team_data["vice_captain"] else "")
        rows.append({
            "Player": f"{p['Player Name']}{marker}",
            "Team": p["Team"],
            "Role": p.get("RoleCode", ""),
            "Tier": p.get("Performance Tier", ""),
            "Credits": p.get("Credits", 0),
            "Est. Pts": p.get("EstimatedPts", 0),
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


# Show My11Circle results
if "my11circle_modes" in st.session_state:
    modes = st.session_state["my11circle_modes"]
    cols = st.columns(3)
    for col, (mode_key, mode_label) in zip(cols, [("safe", "Safe"), ("balanced", "Balanced"), ("aggressive", "Aggressive")]):
        with col:
            st.markdown(f"#### {mode_label}")
            display_team(modes[mode_key], mode_label)

elif "my11circle_team" in st.session_state:
    display_team(st.session_state["my11circle_team"], f"My11Circle ({risk_mode})")

st.markdown("---")

# =====================================================================
# TATA IPL — Transfer-based with smart engine
# =====================================================================
st.subheader("TATA IPL Fantasy — Smart Transfer Engine")
st.caption("160 transfers for 70 matches. Schedule-aware budget. Every transfer must count.")

generate_tata = st.button("Generate TATA IPL Analysis", type="primary", use_container_width=True)

if generate_tata:
    with st.spinner("Running smart transfer analysis..."):
        next_venue = next_match_info.get("VenueName", "") if next_match_info else ""
        tata_result = build_tata_ipl_team(
            players_df, match_info["Team1"], match_info["Team2"],
            match_info["VenueName"], next_venue, playing_xi,
            schedule_df=schedule_df, current_match_num=match_info["Match #"],
        )
        st.session_state["tata_ipl_team"] = tata_result

        # Generate smart transfer recommendations if we have a current team
        if state.get("current_team"):
            current_team_data = []
            for name in state["current_team"]:
                match_data = players_df[players_df["Player Name"] == name]
                if len(match_data) > 0:
                    current_team_data.append(match_data.iloc[0].to_dict())

            all_players = players_df[players_df["IsAvailable"]].to_dict("records")
            transfer_rec = recommend_transfers(
                current_team_data, schedule_df,
                match_info["Match #"], all_players, playing_xi,
            )
            st.session_state["transfer_rec"] = transfer_rec

if "tata_ipl_team" in st.session_state:
    display_team(st.session_state["tata_ipl_team"], "TATA IPL — Optimal XI")

    # Show current squad status with KEEP/SWAP classification
    if "transfer_rec" in st.session_state and state.get("current_team"):
        rec = st.session_state["transfer_rec"]
        st.markdown("---")

        # Budget analysis
        ba = rec.get("budget_analysis", {})
        st.subheader("Smart Transfer Analysis")
        bcol1, bcol2, bcol3, bcol4 = st.columns(4)
        with bcol1:
            st.metric("Smart Budget", f"{ba.get('budget', 0)} transfers")
        with bcol2:
            st.metric("Active Players", rec.get("active_count", 0))
        with bcol3:
            st.metric("Truly Idle", ba.get("truly_idle", 0), help="Idle this match AND next")
        with bcol4:
            st.metric("Remaining", f"{ba.get('remaining', 0)}/160")

        st.info(rec.get("summary", ""))

        # KEEP list
        if rec.get("keep_list"):
            st.markdown("**KEEP (team plays soon — don't waste a transfer):**")
            for k in rec["keep_list"]:
                st.markdown(f"- {k['player']} ({k['team']}) — {k['reason']}")

        # Transfer recommendations
        if rec.get("transfers"):
            st.markdown("**Recommended Transfers:**")
            for i, swap in enumerate(rec["transfers"], 1):
                cls_emoji = {"SWAP": "🔴", "MAYBE": "🟡"}.get(swap.get("out_classification", ""), "🔴")
                with st.expander(
                    f"{cls_emoji} Transfer {i}: {swap['out']} ({swap['out_team']}) → "
                    f"{swap['in_name']} ({swap['in_team']}) | ROI: {swap['roi']:.0f}",
                    expanded=True,
                ):
                    st.markdown(f"**{swap['reason']}**")
                    st.markdown(f"Credit impact: {swap['credit_diff']:+.1f} | Durability: {swap['durability']}/5 upcoming matches")

            st.markdown(f"**Captain (2x):** {rec.get('new_captain', 'N/A')} ({rec.get('captain_pts', 0):.0f} pts)")
            st.markdown(f"**VC (1.5x):** {rec.get('new_vc', 'N/A')} ({rec.get('vc_pts', 0):.0f} pts)")

            if st.button("Apply All Recommended Transfers", type="primary"):
                out_names = [s["out"] for s in rec["transfers"]]
                in_players = [s["in"] for s in rec["transfers"]]
                new_state, errors = apply_transfers(in_players, out_names, rec.get("new_captain"), rec.get("new_vc"))
                if errors:
                    for e in errors:
                        st.error(e)
                else:
                    st.success(f"Applied {len(rec['transfers'])} transfers! {new_state['transfers_remaining']} remaining.")
                    st.rerun()
        else:
            st.success("No transfers needed — squad is well-positioned for this match.")

st.markdown("---")

# --- Save / Set Team ---
if "tata_ipl_team" in st.session_state:
    team_to_save = st.session_state["tata_ipl_team"]
    if team_to_save.get("players") and not state.get("current_team"):
        if st.button("Set as Initial TATA IPL Team (Before Match 1)", type="primary"):
            set_initial_team(team_to_save["players"], team_to_save["captain"], team_to_save["vice_captain"])
            st.success("Initial team set!")
            st.rerun()

# --- Export ---
st.markdown("---")
has_teams = "my11circle_team" in st.session_state or "my11circle_modes" in st.session_state or "tata_ipl_team" in st.session_state
if has_teams and st.button("Export Teams to Excel"):
    import openpyxl

    output_path = Path(__file__).resolve().parent.parent / "data" / "generated_teams"
    output_path.mkdir(parents=True, exist_ok=True)
    match_num = match_info["Match #"]
    filename = output_path / f"Match_{match_num}_Teams.xlsx"

    wb = openpyxl.Workbook()
    teams_to_export = []

    if "my11circle_modes" in st.session_state:
        for mode in ("safe", "balanced", "aggressive"):
            teams_to_export.append((f"My11Circle-{mode.title()}", st.session_state["my11circle_modes"][mode]))
    elif "my11circle_team" in st.session_state:
        teams_to_export.append(("My11Circle", st.session_state["my11circle_team"]))
    if "tata_ipl_team" in st.session_state:
        teams_to_export.append(("TATA IPL", st.session_state["tata_ipl_team"]))

    for sheet_name, team_data in teams_to_export:
        if not team_data.get("players"):
            continue
        ws = wb.create_sheet(sheet_name)
        ws.append(["Player", "Team", "Role", "Tier", "Credits", "Est. Points"])
        for p in team_data["players"]:
            ws.append([p["Player Name"], p["Team"], p.get("RoleCode", ""), p.get("Performance Tier", ""), p.get("Credits", 0), p.get("EstimatedPts", 0)])
        ws.append([])
        ws.append(["Captain", team_data["captain"]])
        ws.append(["Vice-Captain", team_data["vice_captain"]])
        ws.append(["Total Credits", team_data["credits_used"]])
        ws.append(["Strategy", team_data.get("strategy", "")])

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(filename)
    st.success(f"Exported to {filename}")

# --- Differential Picks ---
st.markdown("---")
st.subheader("Differential Picks (High Value, Low Ownership)")
diffs = suggest_differential_picks(players_df, match_info["Team1"], match_info["Team2"], match_info["VenueName"])
if diffs:
    diff_rows = []
    for d in diffs:
        p = d["player"]
        diff_rows.append({
            "Player": p["Player Name"],
            "Team": p["Team"],
            "Role": p.get("RoleCode", ""),
            "Tier": p.get("Performance Tier", ""),
            "Credits": d["credits"],
            "Est. Pts": d["estimated_pts"],
            "Pts/Credit": d["efficiency"],
            "Matchup+": f"{d.get('matchup_bonus', 0):.0f}%",
        })
    st.dataframe(pd.DataFrame(diff_rows), use_container_width=True, hide_index=True)
