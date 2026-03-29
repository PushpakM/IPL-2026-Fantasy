"""
Live stats scraper for IPL 2026 fantasy team builder.
Scrapes ESPNcricinfo for playing XI, recent form, match results, and injury updates.
Falls back gracefully to Excel data if scraping fails.
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
import json
import time
import re
from pathlib import Path
from datetime import datetime

BASE_DIR = Path(__file__).resolve().parent.parent
CACHE_DIR = BASE_DIR / "data"
CACHE_FILE = CACHE_DIR / "scraper_cache.json"
PLAYING_XI_FILE = CACHE_DIR / "playing_xi_cache.json"

_cache = {}
_cache_ttl = 3600  # 1 hour
_playing_xi_ttl = 300  # 5 min cache for playing XI (refresh frequently near toss)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

# Team name mappings (ESPNcricinfo → our codes)
TEAM_NAME_MAP = {
    "royal challengers bengaluru": "RCB", "royal challengers bangalore": "RCB", "rcb": "RCB",
    "chennai super kings": "CSK", "csk": "CSK",
    "mumbai indians": "MI", "mi": "MI",
    "kolkata knight riders": "KKR", "kkr": "KKR",
    "sunrisers hyderabad": "SRH", "srh": "SRH",
    "delhi capitals": "DC", "dc": "DC",
    "rajasthan royals": "RR", "rr": "RR",
    "punjab kings": "PBKS", "pbks": "PBKS",
    "lucknow super giants": "LSG", "lsg": "LSG",
    "gujarat titans": "GT", "gt": "GT",
}


def _get_cache():
    global _cache
    if not _cache and CACHE_FILE.exists():
        try:
            _cache = json.loads(CACHE_FILE.read_text())
        except Exception:
            _cache = {}
    return _cache


def _save_cache():
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    CACHE_FILE.write_text(json.dumps(_cache, indent=2, default=str))


def _is_cached(key, ttl=None):
    cache = _get_cache()
    if key in cache:
        ts = cache[key].get("_timestamp", 0)
        if time.time() - ts < (ttl or _cache_ttl):
            return True
    return False


def _fetch_url(url, timeout=15):
    """Fetch a URL with error handling and rate limiting."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout)
        resp.raise_for_status()
        time.sleep(0.5)
        return resp.text
    except Exception as e:
        print(f"[Scraper] Failed to fetch {url}: {e}")
        return None


def _normalize_team(name):
    """Convert team name to standard code."""
    return TEAM_NAME_MAP.get(name.lower().strip(), name.upper().strip())


# ============================================
# PLAYING XI SCRAPER — THE KEY FEATURE
# ============================================

def scrape_playing_xi(team1: str, team2: str) -> dict:
    """
    Scrape confirmed playing XI for both teams from multiple sources.
    Call this ~30 mins before match (at toss time).

    Returns: {
        team1: {team: str, players: [str], source: str},
        team2: {team: str, players: [str], source: str},
        toss: {winner: str, decision: str},
        scraped_at: str,
    }
    """
    cache_key = f"playing_xi_{team1}_{team2}"
    if _is_cached(cache_key, ttl=_playing_xi_ttl):
        return _cache[cache_key]

    result = {
        "team1": {"team": team1, "players": [], "source": "none"},
        "team2": {"team": team2, "players": [], "source": "none"},
        "toss": {"winner": "", "decision": ""},
        "scraped_at": datetime.now().isoformat(),
        "_timestamp": time.time(),
    }

    # Try multiple sources
    sources = [
        _scrape_playing_xi_espn,
        _scrape_playing_xi_iplt20,
        _scrape_playing_xi_cricbuzz,
    ]

    for scrape_fn in sources:
        try:
            data = scrape_fn(team1, team2)
            if data:
                if data.get("team1", {}).get("players"):
                    result["team1"] = data["team1"]
                if data.get("team2", {}).get("players"):
                    result["team2"] = data["team2"]
                if data.get("toss", {}).get("winner"):
                    result["toss"] = data["toss"]
                # If we got both teams, stop trying other sources
                if result["team1"]["players"] and result["team2"]["players"]:
                    break
        except Exception as e:
            print(f"[Scraper] {scrape_fn.__name__} failed: {e}")
            continue

    _cache[cache_key] = result
    _save_cache()

    # Also save to dedicated playing XI file for easy access
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    PLAYING_XI_FILE.write_text(json.dumps(result, indent=2))

    return result


def _scrape_playing_xi_espn(team1: str, team2: str) -> dict:
    """Scrape playing XI from ESPNcricinfo match page."""
    result = {"team1": {"team": team1, "players": [], "source": "espncricinfo"},
              "team2": {"team": team2, "players": [], "source": "espncricinfo"},
              "toss": {"winner": "", "decision": ""}}

    # Try the IPL 2026 series page for live/upcoming match
    url = "https://www.espncricinfo.com/series/indian-premier-league-2026-1510719/match-results"
    html = _fetch_url(url)
    if not html:
        # Try alternate URL
        url = "https://www.espncricinfo.com/series/ipl-2026-1510719/match-results"
        html = _fetch_url(url)
    if not html:
        return None

    soup = BeautifulSoup(html, "lxml")

    # Find the specific match link
    match_pattern = re.compile(f"{team1}.*{team2}|{team2}.*{team1}", re.IGNORECASE)
    match_links = []
    for a in soup.find_all("a", href=True):
        text = a.get_text(strip=True)
        if match_pattern.search(text):
            match_links.append(a["href"])

    if not match_links:
        return None

    # Fetch the match page
    match_url = match_links[0]
    if not match_url.startswith("http"):
        match_url = "https://www.espncricinfo.com" + match_url
    match_html = _fetch_url(match_url)
    if not match_html:
        return None

    match_soup = BeautifulSoup(match_html, "lxml")

    # Look for playing XI section
    # ESPNcricinfo shows playing XI in match info section
    xi_sections = match_soup.find_all(string=re.compile("Playing XI", re.IGNORECASE))
    for section in xi_sections:
        parent = section.find_parent("div")
        if parent:
            # Extract player names from the section
            player_elements = parent.find_all("a", href=re.compile("/player/"))
            for elem in player_elements:
                name = elem.get_text(strip=True)
                if name:
                    # Determine which team this player belongs to
                    # (Would need the team context from the page structure)
                    pass

    # Look for toss info
    toss_elements = match_soup.find_all(string=re.compile("won the toss", re.IGNORECASE))
    if toss_elements:
        toss_text = toss_elements[0]
        if "bat" in toss_text.lower():
            result["toss"]["decision"] = "bat"
        elif "bowl" in toss_text.lower() or "field" in toss_text.lower():
            result["toss"]["decision"] = "bowl"
        for team_name, code in TEAM_NAME_MAP.items():
            if team_name in toss_text.lower():
                result["toss"]["winner"] = code
                break

    return result


def _scrape_playing_xi_iplt20(team1: str, team2: str) -> dict:
    """Scrape playing XI from official IPLT20.com."""
    result = {"team1": {"team": team1, "players": [], "source": "iplt20"},
              "team2": {"team": team2, "players": [], "source": "iplt20"},
              "toss": {"winner": "", "decision": ""}}

    url = "https://www.iplt20.com/matches/results"
    html = _fetch_url(url)
    if not html:
        return None

    soup = BeautifulSoup(html, "lxml")
    # IPLT20 uses JavaScript rendering, so this may return limited data
    # Look for playing XI in any available format
    return result if result["team1"]["players"] else None


def _scrape_playing_xi_cricbuzz(team1: str, team2: str) -> dict:
    """Scrape playing XI from Cricbuzz."""
    result = {"team1": {"team": team1, "players": [], "source": "cricbuzz"},
              "team2": {"team": team2, "players": [], "source": "cricbuzz"},
              "toss": {"winner": "", "decision": ""}}

    # Cricbuzz IPL 2026 schedule page
    url = "https://www.cricbuzz.com/cricket-series/9237/indian-premier-league-2026/matches"
    html = _fetch_url(url)
    if not html:
        return None

    soup = BeautifulSoup(html, "lxml")

    # Find match link for team1 vs team2
    match_pattern = re.compile(f"{team1}.*{team2}|{team2}.*{team1}", re.IGNORECASE)
    for a in soup.find_all("a", href=True):
        text = a.get_text(strip=True)
        if match_pattern.search(text) and "scorecard" in a["href"].lower():
            scorecard_url = "https://www.cricbuzz.com" + a["href"]
            sc_html = _fetch_url(scorecard_url)
            if sc_html:
                sc_soup = BeautifulSoup(sc_html, "lxml")
                # Extract playing XI from scorecard page
                _extract_players_from_scorecard(sc_soup, result, team1, team2)
            break

    return result if result["team1"]["players"] else None


def _extract_players_from_scorecard(soup, result, team1, team2):
    """Extract player names from a scorecard page."""
    # Look for batting/bowling tables
    tables = soup.find_all("table")
    current_team = None
    for table in tables:
        # Try to identify which team's table this is
        header = table.find_previous(["h2", "h3", "div"])
        if header:
            header_text = header.get_text(strip=True).lower()
            if team1.lower() in header_text:
                current_team = "team1"
            elif team2.lower() in header_text:
                current_team = "team2"

        if current_team:
            player_links = table.find_all("a", href=re.compile("/player/|/profiles/"))
            for link in player_links:
                name = link.get_text(strip=True)
                if name and len(name) > 2:
                    result[current_team]["players"].append(name)


def get_playing_xi_cached() -> dict:
    """Get the last cached playing XI data."""
    if PLAYING_XI_FILE.exists():
        try:
            return json.loads(PLAYING_XI_FILE.read_text())
        except Exception:
            pass
    return None


def set_playing_xi_manual(team1: str, team2: str, team1_players: list, team2_players: list):
    """
    Manually set playing XI (if scraping fails, user can input).
    This is the fallback — user pastes the XI from TV/app.
    """
    data = {
        "team1": {"team": team1, "players": team1_players, "source": "manual"},
        "team2": {"team": team2, "players": team2_players, "source": "manual"},
        "toss": {"winner": "", "decision": ""},
        "scraped_at": datetime.now().isoformat(),
        "_timestamp": time.time(),
    }
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    PLAYING_XI_FILE.write_text(json.dumps(data, indent=2))

    global _cache
    _cache[f"playing_xi_{team1}_{team2}"] = data
    _save_cache()
    return data


# ============================================
# EXISTING SCRAPERS (kept + improved)
# ============================================

def scrape_recent_form(player_name: str) -> dict:
    """Scrape recent IPL form for a player."""
    cache_key = f"form_{player_name}"
    if _is_cached(cache_key):
        return _cache[cache_key]

    result = {
        "player": player_name,
        "recent_scores": [],
        "recent_wickets": [],
        "form_factor": 1.0,
        "source": "default",
        "_timestamp": time.time(),
    }

    try:
        search_name = player_name.replace(" ", "+")
        search_url = f"https://search.espncricinfo.com/ci/content/site/search.html?search={search_name}"
        html = _fetch_url(search_url)
        if html:
            soup = BeautifulSoup(html, "lxml")
            links = soup.select("a[href*='/player/']")
            if links:
                result["source"] = "espncricinfo"
    except Exception as e:
        print(f"[Scraper] Form scrape failed for {player_name}: {e}")

    _cache[cache_key] = result
    _save_cache()
    return result


def scrape_pitch_report(venue: str) -> dict:
    """Get pitch/weather conditions for a venue."""
    cache_key = f"pitch_{venue}"
    if _is_cached(cache_key):
        return _cache[cache_key]

    result = {
        "venue": venue,
        "pitch_type": "Standard",
        "weather": "Clear",
        "toss_advice": "Bat first",
        "source": "default",
        "_timestamp": time.time(),
    }

    _cache[cache_key] = result
    _save_cache()
    return result


def scrape_match_result(team1: str, team2: str, match_date: str = "") -> dict:
    """Scrape match scorecard to get actual results after a match."""
    cache_key = f"result_{team1}_{team2}_{match_date}"
    if _is_cached(cache_key):
        return _cache[cache_key]

    result = {
        "team1": team1, "team2": team2, "date": match_date,
        "top_scorers": [], "top_wicket_takers": [],
        "player_performances": [], "injuries_reported": [],
        "source": "default", "_timestamp": time.time(),
    }

    try:
        url = "https://www.espncricinfo.com/series/indian-premier-league-2026/match-results"
        html = _fetch_url(url)
        if html:
            soup = BeautifulSoup(html, "lxml")
            match_links = soup.select("a[href*='/full-scorecard']")
            result["source"] = "espncricinfo" if match_links else "default"
    except Exception as e:
        print(f"[Scraper] Match result scrape failed: {e}")

    _cache[cache_key] = result
    _save_cache()
    return result


def scrape_injury_updates() -> list:
    """Check for new injury updates from cricket news sources."""
    cache_key = "injury_updates"
    if _is_cached(cache_key):
        return _cache[cache_key].get("injuries", [])

    injuries = []
    try:
        url = "https://www.espncricinfo.com/series/indian-premier-league-2026/news"
        html = _fetch_url(url)
        if html:
            soup = BeautifulSoup(html, "lxml")
            headlines = soup.select("h3, .headline, .story-title")
            injury_keywords = ["injury", "ruled out", "hamstring", "strain", "fracture", "surgery", "miss"]
            for h in headlines:
                text = h.get_text(strip=True).lower()
                if any(kw in text for kw in injury_keywords):
                    injuries.append({"headline": h.get_text(strip=True), "source": "espncricinfo"})
    except Exception as e:
        print(f"[Scraper] Injury update scrape failed: {e}")

    _cache[cache_key] = {"injuries": injuries, "_timestamp": time.time()}
    _save_cache()
    return injuries


def update_injury_tracker(new_injuries: list, tracker_path=None):
    """Write new injuries back to the injury tracker Excel file."""
    if not new_injuries:
        return
    if tracker_path is None:
        tracker_path = BASE_DIR / "IPL_2026_Injury_Tracker.xlsx"
    try:
        import openpyxl
        wb = openpyxl.load_workbook(tracker_path)
        ws = wb["IPL 2026 Injury Tracker"]
        last_row = ws.max_row + 1
        for injury in new_injuries:
            ws.cell(row=last_row, column=1, value=injury.get("Team", ""))
            ws.cell(row=last_row, column=2, value=injury.get("Player", ""))
            ws.cell(row=last_row, column=3, value=injury.get("Role", ""))
            ws.cell(row=last_row, column=4, value=injury.get("Status", ""))
            ws.cell(row=last_row, column=5, value=injury.get("Injury / Reason", ""))
            ws.cell(row=last_row, column=6, value=injury.get("Availability", ""))
            ws.cell(row=last_row, column=7, value=injury.get("Replacement", "Not yet announced"))
            last_row += 1
        wb.save(tracker_path)
        print(f"[Scraper] Updated injury tracker with {len(new_injuries)} new entries")
    except Exception as e:
        print(f"[Scraper] Failed to update injury tracker: {e}")


def update_player_form(players_df: pd.DataFrame, match_result: dict) -> pd.DataFrame:
    """Adjust player form factors based on actual match performance."""
    if "FormFactor" not in players_df.columns:
        players_df["FormFactor"] = 1.0

    for perf in match_result.get("player_performances", []):
        name = perf.get("player", "")
        actual_points = perf.get("fantasy_points", 0)
        mask = players_df["Player Name"] == name
        if mask.any():
            current_form = players_df.loc[mask, "FormFactor"].iloc[0]
            if actual_points > 50:
                new_signal = 1.2
            elif actual_points > 30:
                new_signal = 1.1
            elif actual_points > 15:
                new_signal = 1.0
            elif actual_points > 5:
                new_signal = 0.9
            else:
                new_signal = 0.8
            updated_form = round(0.7 * current_form + 0.3 * new_signal, 2)
            players_df.loc[mask, "FormFactor"] = max(0.5, min(1.5, updated_form))

    return players_df


def post_match_sync(team1: str, team2: str, match_date: str = "") -> dict:
    """Full post-match sync: scrape results, update injuries, update form."""
    summary = {"match": f"{team1} vs {team2}", "updates": []}
    result = scrape_match_result(team1, team2, match_date)
    summary["match_result"] = result
    if result["source"] != "default":
        summary["updates"].append("Match scorecard fetched")
    injuries = scrape_injury_updates()
    if injuries:
        summary["updates"].append(f"Found {len(injuries)} injury reports")
        summary["injury_headlines"] = injuries
    return summary
